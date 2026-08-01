#!/usr/bin/env python3.9
# coding: utf-8

import math
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from matplotlib.patches import FancyArrowPatch
import numpy as np
from openpyxl import Workbook
import os
import pandas as pd
import random
import time
import copy
import datetime
import cProfile
import pstats
from pstats import SortKey
import sys, getopt

from lkh_FDTSP import lkh_solution
import modelo_Gurobi

from funciones_c import factibilidad_c, tiemposArribo_c

def registro_dist(time_t, time_d):

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    # datos = []
    nodos = [i for i in coords]
    for i in nodos:
        for j in nodos:
            if time_t[i][j] != False and time_d[i][j] != False:
                datos = [i, j, time_t[i][j], time_d[i][j]]
        
                try:
                    ws = wb['distancias']
                except KeyError:
                    ws = wb.create_sheet('distancias')
                    ws.append(['i','j', 'camion', 'dron'])
                
                ws.append(datos)

    wb.save('./distancias2.xlsx') 

def lecturaDatos(name_file, instances_path):
    instance = instances_path + '/' + name_file

    file = pd.read_csv(instance)

    string = ''
    param =[]
    for l in name_file:
        if l == '_':
            param.append(string)
            string = ''
        elif l == '.':
            param.append(string)
            string = ''
        else:
            string += l

    n = int(param[1]) # Nr de nodos
    endurance = int(param[2])  # Battery endurance of a drone
    droneSpeed = int(param[3]) 
    truckSpeed = 35 #km/hour

    positions = []
    for i in range(n):
        aux =[]
        x = file['X'][i]
        y = file['Y'][i]
        aux.append(x), aux.append(y)
        positions.append(aux)
    coords= {}
    for i in range(len(positions)):
        coords[i] = positions[i]
    
    coords[n] = coords[0] # Duplicación del deposito

    return n, endurance, droneSpeed, truckSpeed, coords

def tiemposViaje():
    time_t = list()
    time_d = list()
    nodos = [i for i in coords]

    for i in nodos:
        lista_aux_t = list()
        lista_aux_d = list()
        for j in nodos:
            if i != nodos[-1] and j != 0 and i != j:
                if i == 0 and j == nodos[-1]:
                    lista_aux_t.append(False)
                    lista_aux_d.append(False)
                else:
                    lista_aux_t.append( (( abs(float(coords[i][0]) - float(coords[j][0])) 
                                         + abs(float(coords[i][1]) - float(coords[j][1])) )/truckSpeed)*60 ) # min
                    lista_aux_d.append( ((math.sqrt((float(coords[i][0]) - float(coords[j][0]))**2 
                                                  + (float(coords[i][1]) - float(coords[j][1]))**2))/droneSpeed)*60 ) # min
            else: 
                lista_aux_t.append(False)
                lista_aux_d.append(False)

        time_t.append(lista_aux_t)
        time_d.append(lista_aux_d)
        
    return time_t, time_d


def plot_solution_route(
    coords,
    truck_route,
    drone_routes,
    instance_name,
    number_of_drones,
    seed_value,
    objective_value,
    output_dir="route_plots",
):
    """Save a publication-quality plot of a truck--drone solution.

    The truck route is shown with solid arrows. Each drone trip is shown as
    two dashed arrows: launch -> served customer -> recovery. The duplicated
    final depot is displayed at the same coordinates as the initial depot.
    Both PNG and PDF files are generated for use in the manuscript.
    """
    if not truck_route:
        raise ValueError("truck_route cannot be empty")

    os.makedirs(output_dir, exist_ok=True)

    # Customers served by drones and truck, excluding both depot copies.
    drone_customers = {trip[1] for trip in drone_routes}
    final_depot = truck_route[-1]
    depot_nodes = {0, final_depot}
    truck_customers = [
        node for node in truck_route
        if node not in depot_nodes and node not in drone_customers
    ]

    fig, ax = plt.subplots(figsize=(10, 8))

    # Draw the truck route first so nodes remain visible above the arcs.
    for origin, destination in zip(truck_route[:-1], truck_route[1:]):
        x1, y1 = coords[origin]
        x2, y2 = coords[destination]
        arrow = FancyArrowPatch(
            (x1, y1),
            (x2, y2),
            arrowstyle="-|>",
            mutation_scale=12,
            linewidth=1.8,
            linestyle="-",
            color="black",
            shrinkA=5,
            shrinkB=5,
            connectionstyle="arc3,rad=0.0",
            zorder=1,
        )
        ax.add_patch(arrow)

    # Draw each drone trip using a distinct color when possible.
    cmap = plt.get_cmap("tab10")
    for route_index, (launch, customer, recovery) in enumerate(drone_routes):
        route_color = cmap(route_index % 10)
        legs = [(launch, customer), (customer, recovery)]
        for origin, destination in legs:
            x1, y1 = coords[origin]
            x2, y2 = coords[destination]
            arrow = FancyArrowPatch(
                (x1, y1),
                (x2, y2),
                arrowstyle="-|>",
                mutation_scale=11,
                linewidth=1.6,
                linestyle="--",
                color=route_color,
                shrinkA=5,
                shrinkB=5,
                connectionstyle="arc3,rad=0.06",
                zorder=2,
            )
            ax.add_patch(arrow)

    # Plot customer nodes by service mode.
    if truck_customers:
        ax.scatter(
            [coords[node][0] for node in truck_customers],
            [coords[node][1] for node in truck_customers],
            marker="o",
            s=58,
            facecolors="white",
            edgecolors="black",
            linewidths=1.4,
            zorder=4,
        )

    if drone_customers:
        ax.scatter(
            [coords[node][0] for node in sorted(drone_customers)],
            [coords[node][1] for node in sorted(drone_customers)],
            marker="s",
            s=62,
            facecolors="white",
            edgecolors="black",
            linewidths=1.4,
            zorder=4,
        )

    # Initial and final depots share the same physical location.
    depot_x, depot_y = coords[0]
    ax.scatter(
        [depot_x],
        [depot_y],
        marker="*",
        s=220,
        facecolors="white",
        edgecolors="black",
        linewidths=1.6,
        zorder=5,
    )

    # Label all physical customer nodes once; label the depot as 0/n.
    physical_nodes = sorted(node for node in coords if node != final_depot)
    for node in physical_nodes:
        x, y = coords[node]
        label = f"0/{final_depot}" if node == 0 else str(node)
        ax.annotate(
            label,
            (x, y),
            xytext=(5, 5),
            textcoords="offset points",
            fontsize=8,
            zorder=6,
        )

    legend_handles = [
        Line2D([0], [0], color="black", linewidth=1.8, linestyle="-", label="Truck route"),
        Line2D([0], [0], color="black", linewidth=1.6, linestyle="--", label="Drone trip"),
        Line2D([0], [0], marker="o", color="none", markeredgecolor="black", markerfacecolor="white", markersize=7, label="Truck-served customer"),
        Line2D([0], [0], marker="s", color="none", markeredgecolor="black", markerfacecolor="white", markersize=7, label="Drone-served customer"),
        Line2D([0], [0], marker="*", color="none", markeredgecolor="black", markerfacecolor="white", markersize=12, label="Depot"),
    ]
    ax.legend(handles=legend_handles, loc="best", frameon=True)

    ax.set_title(
        f"2SM route: {instance_name} | D={number_of_drones}, seed={seed_value}, "
        f"completion time={objective_value:.2f} min"
    )
    ax.set_xlabel("X coordinate")
    ax.set_ylabel("Y coordinate")
    ax.set_aspect("equal", adjustable="datalim")
    ax.grid(True, linestyle=":", linewidth=0.6, alpha=0.6)
    fig.tight_layout()

    safe_name = os.path.splitext(os.path.basename(instance_name))[0]
    base_path = os.path.join(
        output_dir,
        f"{safe_name}_D{number_of_drones}_seed{seed_value}_2SM_route",
    )
    fig.savefig(base_path + ".png", dpi=300, bbox_inches="tight")
    fig.savefig(base_path + ".pdf", bbox_inches="tight")
    plt.close(fig)

    return base_path + ".png", base_path + ".pdf"


def sol_inicial(ruta_lkh, cytn):

    truckNodes = list(coords.keys())
    truckRoute = lkh_solution(nrNodos, truckNodes, time_t, seed, grid, endurance, droneSpeed, nrDrones, ruta_lkh) 
    truckRoute_lkh = truckRoute.copy()

    if nrNodos - nrDrones < 2:
        asignados = nrNodos - 2 # Asumiendo que la instancia más pequeña es de 5 nodos, la cantidad maxima de clientes atendidos por drones es 3. El camión debe atender al menos a uno, considerando que también visita el depósito.
    else:
        asignados = nrDrones

    droneRoute = []
    droneNodes = []
    lanzamientos = []
    recepciones = []
    truckRoute_idx = [ i for i in range(1, len(truckRoute) - 1)]

    conteo_vuelos = {i: asignados for i in truckRoute} # Vuelos disponibles
    conteo_lanzamientos = {i: 0 for i in truckRoute} # Vuelos comenzados desde los nodos

    tiempos = {}
    
    tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, tpoArribo = tiempos)

    cant_rutas = len(droneRoute)
    t_ini = time.time() # tiempo de inicio

    for n in range(1, len(truckRoute_lkh) - 1):

        flag = False
        if len(truckRoute_idx) == 0:
            break
        i = truckRoute[random.choice(truckRoute_idx)] # Cliente seleccionado aleatoriamente
        
        if time_t[truckRoute[truckRoute.index(i) - 1]][i] + serv_t + time_t[i][truckRoute[truckRoute.index(i) + 1]] < time_t[truckRoute[truckRoute.index(i) - 1]][truckRoute[truckRoute.index(i) + 1]] + se: # Se evita incrementar el tpo de la ruta del camion
            truckRoute_idx.remove(truckRoute.index(i))
            continue 
        else:
            if i not in droneNodes and i not in lanzamientos and i not in recepciones: # El nodo seleccionado no puede ser uno ya atendido por drones, ni puede ser nodo de lanzamiento o recepción.
                for j in truckRoute_lkh[:-2]: # Nodo de lanzamiento, Depósito inicial + clientes 
                    if j not in droneNodes and j not in lanzamientos and j not in recepciones:
                        if j != i and conteo_vuelos[j] > 0: 
                            for k in truckRoute_lkh[truckRoute_lkh.index(j) + 1:]: # Nodo de recepción, Clientes (a partir del nodo anteriormente seleccionado) + depósito final
                                if k not in droneNodes and k not in lanzamientos and k not in recepciones:
                                    if j == 0 and k == 0:
                                        continue
                                    elif j == 0 and k == truckRoute_lkh[-1]:
                                        continue
                                    else:
                                        flag2 = True 
                                        for h in truckRoute[truckRoute.index(j) + 1: truckRoute.index(k) + 1]:
                                            if conteo_vuelos[h] <= 0:
                                                flag2 = False
                                                break
                                        if k != i and flag2 == True: 
                                            t_se = (conteo_lanzamientos[j] + 1)*se # Tiempo de set up por cada lanzamiento

                                            espera = 0 # Tiempo de espera del dron al camión en el nodo k
                                            espera_camion = 0 # Tiempo de espera del camión al dron en el nodo k
                                            
                                            if j == k: # Loop del dron en un mismo nodo
                                                espera_camion += t_se + time_d[j][i] + serv_d + time_d[i][k]
                                            elif j != k:
                                                if tiempos[j] + serv_t + time_d[j][i] + serv_d + time_d[i][k] < tiempos[k] + t_se: # El camión llega después, el dron debe esperar # No se incluye t_se
                                                    espera += (tiempos[k] + t_se) - (tiempos[j] + serv_t + time_d[j][i] + serv_d + time_d[i][k])

                                                elif tiempos[j] + t_se + serv_t + time_d[j][i] + serv_d + time_d[i][k] > tiempos[k] + t_se: # El dron llega después, el camión debe esperar
                                                    espera_camion += (tiempos[j] + t_se + serv_t + time_d[j][i] + serv_d + time_d[i][k]) - (tiempos[k] + t_se)

                                            if time_d[j][i] + serv_d + time_d[i][k] + espera <= endurance and espera_camion == 0: # El vuelo del dron no excede el límite de la batería
                                                
                                                ok = True
                                                for vuelo in droneRoute: # Verificando que los vuelos no se intersecten
                                                    if truckRoute_lkh.index(j) < truckRoute_lkh.index(vuelo[0]) and truckRoute_lkh.index(k) < truckRoute_lkh.index(vuelo[0]): # El recorrido comienza antes del comienzo de la ruta "vuelo" - termina antes del comienzo de la ruta "vuelo" # Caso 1
                                                        continue # No afecta
                                                
                                                    elif truckRoute_lkh.index(j) < truckRoute_lkh.index(vuelo[0]) and truckRoute_lkh.index(k) > truckRoute_lkh.index(vuelo[0]) and truckRoute_lkh.index(k) < truckRoute_lkh.index(vuelo[2]): # El recorrido comienza antes del comienzo de la ruta "vuelo" - termina despues del comienzo y antes del termino de la ruta "vuelo" # Caso 2
                                                        if espera_camion != 0:
                                                            ok = False
                                                            break # Si afecta por la posible espera del camión, se podría incluir si es que no existiera espera
                                                        elif espera_camion == 0:
                                                            continue                     

                                                    elif truckRoute_lkh.index(j) < truckRoute_lkh.index(vuelo[0]) and truckRoute_lkh.index(k) > truckRoute_lkh.index(vuelo[2]): # El recorrido comienza antes del comienzo de la ruta "vuelo" y termina despues del termino de la ruta "vuelo" # Caso 3
                                                        continue # aqui la ruta j,i,k se ve afectada por "vuelo", si los tiempos de espera ya estan considerados no afectaría      

                                                    elif truckRoute_lkh.index(j) > truckRoute_lkh.index(vuelo[0]) and truckRoute_lkh.index(k) < truckRoute_lkh.index(vuelo[2]): # El recorrido comienza despues del comienzo de la ruta "vuelo" y termina antes del termino de la ruta "vuelo"
                                                        ok = False
                                                        break # Si afecta por se y la posible espera del camión. Se podría incluir si es que no existiera espera y si el se no afecta el "vuelo"

                                                    elif truckRoute_lkh.index(j) > truckRoute_lkh.index(vuelo[0]) and truckRoute_lkh.index(j) < truckRoute_lkh.index(vuelo[2]) and truckRoute_lkh.index(k) > truckRoute_lkh.index(vuelo[2]): # El recorrido comienza despues del comienzo y antes del termino de la ruta "vuelo" - termina despues del termino de la ruta "vuelo"
                                                        ok = False
                                                        break # Si afecta por el se y se ve afectado por la espera del "vuelo". Si los tiempos de espera ya estan considerados no afectaría. Se podría incluir si es que el se no afecta el "vuelo"

                                                    elif truckRoute_lkh.index(j) > truckRoute_lkh.index(vuelo[2]) and truckRoute_lkh.index(k) > truckRoute_lkh.index(vuelo[2]): # El recorrido comienza despues del termino de la ruta "vuelo" - termina despues del termino de la ruta "vuelo"
                                                        continue # No afecta
                                                        
                                                if ok == True:
                                                    
                                                    truckRoute_idx.remove(truckRoute.index(i)) # Los indices estan actualizados, y se removión el indice del nodo atendido. Falta remover los indices de los nodos de lanzamiento, entremedio y recepcion.
                                                    a = []
                                                    for l in range(len(truckRoute_idx)):
                                                        if truckRoute.index(j) <= truckRoute_idx[l] and truckRoute_idx[l] <= truckRoute.index(k):
                                                            a.append(truckRoute_idx[l])

                                                    for l in a:
                                                        truckRoute_idx.remove(l)

                                                    for l in range(len(truckRoute_idx)):
                                                        if truckRoute_idx[l] > truckRoute.index(i):
                                                            truckRoute_idx[l] -= 1

                                                    droneRoute.append([j,i,k])
                                                    droneNodes.append(i)
                                                    truckRoute.remove(i)
                                                    flag = True
                                                    lanzamientos.append(j)
                                                    recepciones.append(k)

                                                    conteo_lanzamientos[j] += 1
                                                    conteo_vuelos.pop(i)
                                                    conteo_vuelos[j] -= 1
                                                    if j != k: # Ajustando balance de drones disponibles
                                                        for l in truckRoute_lkh:
                                                            if l not in droneNodes:
                                                                if truckRoute_lkh.index(l) > truckRoute_lkh.index(j) and truckRoute_lkh.index(l) < truckRoute_lkh.index(k):
                                                                    conteo_vuelos[l] -=1
                                                        
                                                        if conteo_vuelos[k] < asignados:
                                                            conteo_vuelos[k] += 1

                                                    break # Este break detiene el ciclo del de nodo de recepcion
                            if flag == True:
                                break # Este break detiene el ciclo del nodo de lanzamiento  

        if len(droneRoute) != 0:
            tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, droneRoute, tpoArribo = tiempos)
        
        if len(droneRoute) == cant_rutas or len(truckRoute_idx) == 0 or time.time() - t_ini > 2: # No se encontraron más rutas
            break
    
    if cytn == 0:
        infactibilidades = factibilidad_func(truckRoute, droneRoute, tiempos)
    elif cytn == 1:
        infactibilidades = factibilidad_c(time_d, serv_t, serv_d, se, nrDrones, endurance, truckRoute, droneRoute, tiempos)

    if infactibilidades != 0: # Se continuará con la ruta generada en LKH.
        tiempos = {}

        tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute_lkh, tpoArribo = tiempos)

        truckRoute = truckRoute_lkh.copy()
        droneRoute = []
        droneNodes = [] 
        infactibilidades = 0 

    return truckRoute, droneRoute, droneNodes, tiempos, infactibilidades

def dronVecinoAleatorio(truckRoute, droneNodes): # Se construye la ruta de los drones asignando aleatoriamente los nodos de lanzamiento y recepcion segun un nodo asignado a los drones.
    droneRoute = []
    for node in droneNodes:
        while True:
            launchIndex =  random. randint(0, len(truckRoute) - 2) 
            if launchIndex == 0:
                retrievalIndex = random. randint(1, len(truckRoute) - 2)

            elif launchIndex != 0:
                retrievalIndex = random. randint(launchIndex, len(truckRoute) - 1)

            launch = truckRoute[launchIndex]
            retrieval = truckRoute[retrievalIndex]
            if launch not in droneNodes and retrieval not in droneNodes:
                trip = [launch, node, retrieval]
                break
        droneRoute.append(trip)

    return droneRoute

def perturbacionDron2(truckRoute, droneRoute, indice_ruta): # Se seleccionan dos nodos aleatorios de la ruta del camion, si son distintos a los nodos de lanzamiento y recepcion se intercambian estos (1 o los 2 con un probabilidad)
    l_t = len(truckRoute)
    cambio = random.randint(0,1)
    if l_t > 3:
        route = droneRoute[indice_ruta]
        
        if cambio == 0:
            
            if route[0] == 0 and route[2] != truckRoute[-1]:
                i = random.randint(1,truckRoute.index(route[2]))
            
            elif route[0] != 0 and route[2] == truckRoute[-1]:
                i = random.randint(1,l_t - 2)
                while truckRoute[i] == route[0]:
                    i = random.randint(1,l_t - 2)

            elif route[0] != 0 and route[2] != truckRoute[-1]:
                i = random.randint(0,truckRoute.index(route[2]))
                while truckRoute[i] == route[0]:
                    i = random.randint(0,truckRoute.index(route[2]))

            if time_d[truckRoute[i]][ route[1]] + serv_d + time_d[route[1]][ route[2]] <= endurance: # El viaje del dron, sin considerar tiempo de espera, es factible
                droneRoute[indice_ruta][0] = truckRoute[i]
        
        elif cambio == 1:
            
            if route[2] == truckRoute[-1] and route[0] != 0:
                k = random.randint(truckRoute.index(route[0]),l_t - 2)
            
            elif route[2] != truckRoute[-1] and route[0] == 0:
                k = random.randint(1,l_t - 2)
                while truckRoute[k] == route[2]:
                    k = random.randint(1,l_t - 2)
            
            elif route[2] != truckRoute[-1] and route[0] != 0:
                k = random.randint(truckRoute.index(route[0]),l_t - 1)
                while truckRoute[k] == route[2]:
                    k = random.randint(truckRoute.index(route[0]),l_t - 1)

            if time_d[route[0]][ route[1]] + serv_d + time_d[route[1]][ truckRoute[k]] <= endurance: # El viaje del dron, sin considerar tiempo de espera, es factible
                droneRoute[indice_ruta][2] = truckRoute[k]

def reducir_op2(truckRoute, droneNodes, droneRoute, nrCambios = 1): # Se selecciona un nodo atendido por el camión y es pasado a algún dron.
    if len(truckRoute) - nrCambios > 3 and len(truckRoute) > 3:

        nodos = [] # Nodos de lanzamiento y recepción
        rutas = []
        for i in droneRoute:
            if i[0] not in nodos:
                nodos.append(i[0])
            if i[2] not in nodos:
                nodos.append(i[2])
            
            rutas.append((truckRoute.index(i[0]), truckRoute.index(i[2])))

        nodos_libres = [i for i in truckRoute if i not in nodos and i != 0 and i != truckRoute[-1]] # Posibles nodos a intercambiar

        nodos_intermedios = []
        flag = False
        for i in nodos_libres:
            for j in rutas:
                if j[0] <  truckRoute.index(i) < j[1]:
                    nodos_intermedios.append(i)
                    nodos_libres.remove(i)
                    flag = True 
                    break
            if flag == True:
                break
        
        if len(nodos_libres) != 0:
            n = random.choice(nodos_libres)
            nodos_libres.remove(n)
            flag2 = False
            
            if time_t[truckRoute.index(n) - 1][ truckRoute.index(n)] + serv_t + time_t[truckRoute.index(n)][ truckRoute.index(n) + 1] < time_t[truckRoute.index(n) - 1][ truckRoute.index(n) + 1]:
                flag = False

            while flag == False: # se busca que el recorrido nuevo no se mayor que el previo
                if len(nodos_libres) != 0:
                    n = random.choice(nodos_libres)
                    nodos_libres.remove(n)
                    if time_t[truckRoute.index(n) - 1][ truckRoute.index(n)] + serv_t + time_t[truckRoute.index(n)][ truckRoute.index(n) + 1] >= time_t[truckRoute.index(n) - 1][ truckRoute.index(n) + 1]:
                        break
                    
                elif len(nodos_intermedios) != 0:
                    n = random.choice(nodos_intermedios)
                    nodos_intermedios.remove(n)
                    if time_t[truckRoute.index(n) - 1][ truckRoute.index(n)] + serv_t + time_t[truckRoute.index(n)][ truckRoute.index(n) + 1] >= time_t[truckRoute.index(n) - 1][ truckRoute.index(n) + 1]:
                        break
                
                if len(nodos_libres) == 0 and len(nodos_intermedios) == 0:
                    flag2 = True
                    break
            
            if flag2 == True:
                n = random.choice([i for i in truckRoute if i not in nodos and i != 0 and i != truckRoute[-1]])

            droneNodes.append(n)
            truckRoute.remove(n)

def bl_reducir_op(truckRoute, droneNodes, droneRoute): # Se hace una BL a los nodos asignados a ser atendidos por los drones que no tengan ruta asignada.
    
    if len(droneNodes) != len(droneRoute):
        
        if nrNodos - nrDrones < 2: # Evitar problemas con el nr de drones
            asignados = nrNodos - 2
        else:
            asignados = nrDrones

        lanzamientos = []
        recepciones = []

        conteo_vuelos = {i: asignados for i in truckRoute}
        conteo_lanzamientos = {i: 0 for i in truckRoute}

        for i in droneRoute:
            if i[0] not in lanzamientos:
                lanzamientos.append(i[0])
            if i[2] not in recepciones:
                recepciones.append(i[2])
            
            conteo_lanzamientos[i[0]] += 1

        indices = [(truckRoute.index(i[0]), truckRoute.index(i[2])) for i in droneRoute]
        for i in indices:
            if i[0] != i[1]:
                for j in range(i[0], i[1]):
                    conteo_vuelos[truckRoute[j]] -= 1 # Balance actual de drones disponibles
            
        nodos = []
        for i in droneNodes[::-1]:
            flag = False
            for j in droneRoute:
                if j[1] == i:
                    flag = True
                    break # El nodo ya tiene una ruta de dron asignada
            if flag == False:
                nodos.append(i)
            
        for i in nodos:
            flag = False

            tiempos = {}
            tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, droneRoute, tpoArribo = tiempos)
            
            for j in truckRoute[:-1]: # Nodo de lanzamiento, Depósito inicial -> clientes 
                if j not in recepciones: 
                    if j != i and conteo_vuelos[j] > 0: # Hay drones disponibles para lanzar
                        for k in truckRoute[truckRoute.index(j):]: # Nodo de recepción, Clientes (a partir del nodo de lanzamiento seleccionado) -> depósito final
                            if k not in lanzamientos and k not in recepciones:
                                if j == 0 and k == 0:
                                    continue
                                elif j == 0 and k == truckRoute[-1]:
                                    continue
                                else:
                                    if k != i and conteo_vuelos[k] > 0:

                                        t_se = (conteo_lanzamientos[j] + 1)*se # Tiempo de set up por cada lanzamiento

                                        espera = 0 # Tiempo de espera del dron al camión en el nodo k
                                        espera_camion = 0 # Tiempo de espera del camión al dron en el nodo k
                                        
                                        if j == k: # Loop del dron en un mismo nodo
                                            espera_camion += t_se + time_d[j][i] + serv_d + time_d[i][k]
                                        elif j != k:
                                            if tiempos[j] + serv_t + time_d[j][i] + serv_d + time_d[i][k] < tiempos[k] + t_se: # El camión llega después, el dron debe esperar
                                                espera += (tiempos[k] + t_se) - (tiempos[j] + serv_t + time_d[j][i] + serv_d + time_d[i][k])

                                            elif tiempos[j] + t_se + serv_t + time_d[j][i] + serv_d + time_d[i][k] > tiempos[k] + t_se: # El dron llega después, el camión debe esperar
                                                espera_camion += (tiempos[j] + t_se + serv_t + time_d[j][i] + serv_d + time_d[i][k]) - (tiempos[k] + t_se)

                                        if time_d[j][i] + serv_d + time_d[i][k] + espera <= endurance and espera_camion == 0:  #: # El vuelo del dron no excede el límite de la batería y el camion no debe esperar

                                            ok = True
                                            for vuelo in droneRoute:
                                                if truckRoute.index(j) < truckRoute.index(vuelo[0]) and truckRoute.index(k) < truckRoute.index(vuelo[0]): # El recorrido comienza antes del comienzo de la ruta "vuelo" - termina antes del comienzo de la ruta "vuelo" # Caso 1
                                                    continue
                                                elif truckRoute.index(j) < truckRoute.index(vuelo[0]) and truckRoute.index(k) > truckRoute.index(vuelo[0]) and truckRoute.index(k) < truckRoute.index(vuelo[2]): # El recorrido comienza antes del comienzo de la ruta "vuelo" - termina despues del comienzo y antes del termino de la ruta "vuelo" # Caso 2
                                                    continue
                                                elif truckRoute.index(j) < truckRoute.index(vuelo[0]) and truckRoute.index(k) > truckRoute.index(vuelo[2]): # El recorrido comienza antes del comienzo de la ruta "vuelo" y termina despues del termino de la ruta "vuelo" # Caso 3
                                                    continue 
                                                elif truckRoute.index(j) > truckRoute.index(vuelo[0]) and truckRoute.index(k) < truckRoute.index(vuelo[2]): # El recorrido comienza despues del comienzo de la ruta "vuelo" y termina antes del termino de la ruta "vuelo"
                                                    ok = False
                                                    break # Se ve afectado por el timepo de set up y la posible espera del camión que se genera.
                                                elif truckRoute.index(j) > truckRoute.index(vuelo[0]) and truckRoute.index(j) < truckRoute.index(vuelo[2]) and truckRoute.index(k) > truckRoute.index(vuelo[2]): # El recorrido comienza despues del comienzo y antes del termino de la ruta "vuelo" - termina despues del termino de la ruta "vuelo"
                                                    ok = False
                                                    break # Se ve afectado por el timepo de set up y la posible espera del camión que se genera.
                                                elif truckRoute.index(j) > truckRoute.index(vuelo[2]) and truckRoute.index(k) > truckRoute.index(vuelo[2]): # El recorrido comienza despues del termino de la ruta "vuelo" - termina despues del termino de la ruta "vuelo"
                                                    continue
                                            
                                            if ok == True:
                                                droneRoute.append([j,i,k])
                                                flag = True
                                                lanzamientos.append(j)
                                                recepciones.append(k)
                                                conteo_lanzamientos[j] += 1
                                                conteo_vuelos[j] -= 1
                                                if j != k:
                                                    for l in truckRoute:
                                                        if truckRoute.index(l) > truckRoute.index(j) and truckRoute.index(l) < truckRoute.index(k):
                                                            conteo_vuelos[l] -=1
                                                    
                                                    if conteo_vuelos[k] < asignados:
                                                        conteo_vuelos[k] += 1
                                                
                                                nodos.remove(i)
                                                break # break para el loop de nodos de recepcion
                        if flag == True:
                            break # break par el loop de nodos de lanzamiento

    return nodos # La lista de nodos sin atender como retorno

def insertar(truckRoute, droneNodes, droneRoute, retorno = []): # Se toma un nodo de la ruta de los drones y se inserta en la ruta del camion
    
    pos_disponibles = [i for i in range(1, len(truckRoute))]

    for vuelo in droneRoute:
        for i in range(truckRoute.index(vuelo[0]), truckRoute.index(vuelo[2]) + 1):
            if i in pos_disponibles:
                pos_disponibles.remove(i)

    if len(retorno) == 0:
        if len(droneNodes) != 0 and len(truckRoute) != nrNodos + 1:

            if len(pos_disponibles) != 0:
                vuelo = droneRoute[random.randint(0, len(droneRoute) - 1)] # Solo se selcciona 1 nodo
                pos_en_ruta = random.choice(pos_disponibles)

                truckRoute.insert(pos_en_ruta, vuelo[1])
                droneNodes.remove(vuelo[1])
                droneRoute.remove(vuelo)

    else:
        for i in retorno:
            if len(pos_disponibles) == 0:
                break

            pos_en_ruta = random.choice(pos_disponibles)
            truckRoute.insert(pos_en_ruta, i)
            droneNodes.remove(i)
            retorno.remove(i)
            pos_disponibles.remove(pos_en_ruta)

def perturbacionTSP_op(truckRoute): # O(1) # Se seleccionan dos nodos aleatoriamnete y se intercambian sus posicinoes 
    n = len(truckRoute)

    if n > 3:
        index1 = random.randint(1, n - 3)
        index2 = random.randint(index1 + 1, n - 2)

        node1 = truckRoute[index1]
        node2 = truckRoute[index2]

        truckRoute[index1] = node2
        truckRoute[index2] = node1

def DosOpt_op(truckRoute): # O(n^2)
    n = len(truckRoute)
    flag = True
    count = 0
    for i in range(n - 2):
        for j in range(i + 1, n - 1):
            newCost = time_t[truckRoute[i]][ truckRoute[j]] + time_t[truckRoute[i + 1]][ truckRoute[j + 1]] - time_t[truckRoute[i]][ truckRoute[i + 1]] - time_t[truckRoute[j]][ truckRoute[j + 1]]
            if newCost < 0:
                min_i, min_j = i, j
                count += 1
                if count == 1: # terminar a la primera mejora
                    flag = False

        if flag == False:
            break

    if count > 0:
        truckRoute[min_i + 1 : min_j + 1] = truckRoute[min_i + 1 : min_j + 1][::-1]

def construccion_dron(truckRoute, droneRoute, droneNodes, tiempos, nodos):
    if nrNodos - nrDrones < 2:
        asignados = nrNodos - 2
    else:
        asignados = nrDrones

    droneRoute.clear()
    lanzamientos = []
    recepciones = []

    conteo_vuelos = {i: nrDrones for i in truckRoute}
    conteo_lanzamientos = {i: 0 for i in truckRoute}
    
    for dron in range(asignados): # Por cada dron asignado
        cant_rutas = len(droneRoute)
        flag = False

        for i in droneNodes: # clientes
            if i not in nodos:
                for j in truckRoute[1:-2]: # Nodo de lanzamiento, cliente inicial + clientes 
                    if j not in recepciones and conteo_vuelos[j] > 0:
                        for k in truckRoute[truckRoute.index(j):-1]: # Nodo de recepción, Clientes (a partir del nodo anteriormente seleccionado) + depósito final
                            flag2 = True 
                            for h in truckRoute[truckRoute.index(j) + 1: truckRoute.index(k) + 1]:
                                if conteo_vuelos[h] <= 0:
                                    flag2 = False
                                    break
                            if k not in lanzamientos and k not in recepciones and flag2 == True:
                                if j == 0 and k == 0:
                                    continue
                                elif j == 0 and k == truckRoute[-1]:
                                    continue
                                else:
                                    t_se = (conteo_lanzamientos[j] + 1)*se # Tiempo de set up por cada lanzamiento

                                    espera = 0 # Tiempo de espera del dron al camión en el nodo k
                                    espera_camion = 0 # Tiempo de espera del camión al dron en el nodo k
                                    
                                    if j == k: 
                                        espera_camion += t_se + time_d[j][i] + serv_d + time_d[i][k]
                                    elif j != k:
                                        if tiempos[j] + serv_t + time_d[j][i] + serv_d + time_d[i][k] < tiempos[k] + t_se: # El camión llega después, el dron debe esperar
                                            espera += (tiempos[k] + t_se) - (tiempos[j] + serv_t + time_d[j][i] + serv_d + time_d[i][k])

                                        elif tiempos[j] + t_se + serv_t + time_d[j][i] + serv_d + time_d[i][k] > tiempos[k] + t_se: # El dron llega después, el camión debe esperar
                                            espera_camion += (tiempos[j] + t_se + serv_t + time_d[j][i] + serv_d + time_d[i][k]) - (tiempos[k] + t_se)

                                    if time_d[j][i] + serv_d + time_d[i][k] + espera <= endurance and espera_camion == 0: # El vuelo del dron no excede el límite de la batería y el camion no debe esperar

                                        ok = True
                                        for vuelo in droneRoute: # Verificando que los vuelos no se intersecten
                                            if truckRoute.index(j) < truckRoute.index(vuelo[0]) and truckRoute.index(k) < truckRoute.index(vuelo[0]): # El recorrido comienza antes del comienzo de la ruta "vuelo" - termina antes del comienzo de la ruta "vuelo" # Caso 1
                                                continue 

                                            elif truckRoute.index(j) < truckRoute.index(vuelo[0]) and truckRoute.index(k) > truckRoute.index(vuelo[0]) and truckRoute.index(k) < truckRoute.index(vuelo[2]): # El recorrido comienza antes del comienzo de la ruta "vuelo" - termina despues del comienzo y antes del termino de la ruta "vuelo" # Caso 2
                                                if espera_camion != 0:
                                                    ok = False
                                                    break
                                                elif espera_camion == 0:
                                                    continue

                                            elif truckRoute.index(j) < truckRoute.index(vuelo[0]) and truckRoute.index(k) > truckRoute.index(vuelo[2]): # El recorrido comienza antes del comienzo de la ruta "vuelo" y termina despues del termino de la ruta "vuelo" # Caso 3
                                                continue

                                            elif truckRoute.index(j) > truckRoute.index(vuelo[0]) and truckRoute.index(k) < truckRoute.index(vuelo[2]): # El recorrido comienza despues del comienzo de la ruta "vuelo" y termina antes del termino de la ruta "vuelo"
                                                ok = False
                                                break

                                            elif truckRoute.index(j) > truckRoute.index(vuelo[0]) and truckRoute.index(j) < truckRoute.index(vuelo[2]) and truckRoute.index(k) > truckRoute.index(vuelo[2]): # El recorrido comienza despues del comienzo y antes del termino de la ruta "vuelo" - termina despues del termino de la ruta "vuelo"
                                                ok = False
                                                break 

                                            elif truckRoute.index(j) > truckRoute.index(vuelo[2]) and truckRoute.index(k) > truckRoute.index(vuelo[2]): # El recorrido comienza despues del termino de la ruta "vuelo" - termina despues del termino de la ruta "vuelo"
                                                continue

                                        if ok == True:
                                            droneRoute.append([j,i,k])
                                            nodos.append(i)
                                            flag = True
                                            lanzamientos.append(j)
                                            recepciones.append(k)

                                            conteo_lanzamientos[j] += 1
                                            conteo_vuelos[j] -= 1
                                            if j != k: # Ajustando balance de drones disponibles
                                                for l in truckRoute:
                                                    if l not in droneNodes:
                                                        if truckRoute.index(l) > truckRoute.index(j) and truckRoute.index(l) < truckRoute.index(k):
                                                            conteo_vuelos[l] -=1
                                                
                                                if conteo_vuelos[k] < asignados:
                                                    conteo_vuelos[k] += 1
                                            break
                        if flag == True:
                            break
                if flag == True: 
                    break        

        if len(droneRoute) != 0:
            tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, droneRoute, tpoArribo = tiempos)
                
        if len(droneRoute) == cant_rutas: # No se encontraron más rutas
            break

def RCST(truckRoute, droneRoute, droneNodes): # RCST: Reduction of Customers Served by Truck
    
    reducir_op2(truckRoute, droneNodes, droneRoute)

    if len(droneNodes) != len(droneRoute): # Se agregó un nodo a cacandidateDroneNodes
        
        retorno = bl_reducir_op(truckRoute, droneNodes, droneRoute)
       
        if len(retorno) != 0:
            ICST(truckRoute, droneRoute, droneNodes, retorno)

def ICST(truckRoute, droneRoute, droneNodes, retorno = []): # ICST: Increase in Customers Served by Truck

    len_t = len(truckRoute)
    insertar(truckRoute, droneNodes, droneRoute, retorno)
    if len_t == len(truckRoute) or len(retorno) != 0:  # No se inserto un nuevo nodo en la ruta del camion
        TRA(truckRoute, droneRoute, droneNodes)

def TRA(truckRoute, droneRoute, droneNodes): # TRA: Truck Route Alteration
    perturbacionTSP_op(truckRoute) # Perturbación
    DosOpt_op(truckRoute) # Busqueda Local

    # Verificar si con el cambio de la ruta camion permite a las rutas de drones seguir factibles. Si no lo son, iniciar reconstrucción.
    try:
        tiempos ={}
        tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, droneRoute, tpoArribo = tiempos)
        fact = factibilidad_c(time_d, serv_t, serv_d, se, nrDrones, endurance ,truckRoute, droneRoute, tiempos)
        
    except KeyError:
        fact = 999

    if fact > 0 or len(droneNodes) != len(droneRoute): # Es necesario reconstruir todos los vuelos de los drones
        # Tiempos de arribo de la ruta del camion
        tiempos = {}
        tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, tpoArribo = tiempos)

        nodos = [] # Nodos con rutas asignadas
        construccion_dron(truckRoute, droneRoute, droneNodes, tiempos, nodos)

        if len(droneNodes) != len(droneRoute):
            retorno = droneNodes.copy()
            for i in nodos:
                retorno.remove(i)
            ICST(truckRoute, droneRoute, droneNodes, retorno) # Existen nodos sin poder ser atendidos por los drones, por lo tanto se inserta en la ruta del camion

def tiemposArribo_func(truckRoute, droneRoute='', tpoArribo = '', metrica = ''):
    if metrica == 'SI':
        waiting_time = 0
    # Tiempos de arribo solo para ruta del camion (TSP)
    t = 0
    tpoArribo[0] = t
    for i in range(len(truckRoute) - 1):
        t = time_t[truckRoute[i]][truckRoute[i + 1]]
        
        if i == 0:
            tpoArribo[truckRoute[i + 1]] = tpoArribo[truckRoute[i]] + t
        else: 
            tpoArribo[truckRoute[i + 1]] = tpoArribo[truckRoute[i]] + t + serv_t

        # tpoArribo = tiemposDict(truckRoute)


    if type(droneRoute) != str and len(droneRoute) != 0:

        for i in truckRoute:
            contador_lanzamientos = 0
            tpoRetorno = -1

            for j in droneRoute:
                if i == j[0]:
                    contador_lanzamientos += 1
                    if i == j[2]:
                        if tpoRetorno < time_d[j[0]][j[1]] + serv_d + time_d[j[1]][j[2]]:
                            tpoRetorno = time_d[j[0]][ j[1]] + serv_d + time_d[j[1]][j[2]] # Finalmente se almacena el viaje loop más largo.
            
            # Actualización del tiempo de arribo de los siguientes nodos visitados por el camióm, si es que existe un vuelo tipo "loop"
            if tpoRetorno != -1:
                if metrica == 'SI':
                    waiting_time += tpoRetorno
                index_l = truckRoute.index(i) # Índice del nodo de lanzamiento 
                for k in range(int(index_l) + 1, len(truckRoute)):
                    tpoArribo[truckRoute[k]] += tpoRetorno

            # Actualización del tiempo de arribo de los nodos visitados por los drones y nodos siguientes del camión
            for j in droneRoute:
                if i == j[0]:
                    # Actualizacion del arribo de los nodos visitados por el camión
                    index_l = truckRoute.index(i) # índice del lanzamiento

                    for k in range(int(index_l) + 1, len(truckRoute)):
                        tpoArribo[truckRoute[k]] += se # Se suma el tiempo de set up a los arribo de los nodos que son proximamente visitados

                    # Actualizacion del arribo del nodo visitado por el dron    
                    tpoActual = tpoArribo[i]
                    if j[0] == 0:
                        tpoActual += time_d[j[0]][j[1]] + (se * contador_lanzamientos)
                    elif j[0] != 0:
                        tpoActual += time_d[j[0]][j[1]] + (se * contador_lanzamientos) + serv_t 
                    tpoArribo[j[1]] = tpoActual
                
                elif i != j[0] and i == j[2] and i != truckRoute[-1]:
                        tpoActual = tpoArribo[j[1]] + serv_d + time_d[j[1]][j[2]]

                        if tpoActual > tpoArribo[j[2]]: # El camión debe esperar al dron
                            # Actualizacion de los arribos del camión, según la espera al dron. 
                            espera = tpoActual - tpoArribo[j[2]]
                            if metrica == 'SI':
                                waiting_time += espera
                            index_r = truckRoute.index(j[2]) #Índice del nodo de recepción.
                            for k in range(int(index_r), len(truckRoute)):
                                tpoArribo[truckRoute[k]] += espera
                                for n in droneRoute:
                                    if n[0] == truckRoute[k] and n[1] in tpoArribo:
                                        tpoArribo[n[1]] += espera
    if metrica == "SI":
        return waiting_time

def factibilidad_func(truckRoute, droneRoute='', tiemposArribo = ''): # Se evaluan distintos tipos de infactibilidad. Se pueden identificar cada uno, pero se penaliza de la misma forma para cada tipo.
    if type(droneRoute) != str:
            
        conteoL = {i:0 for i in truckRoute} # Conteo de lanzamientos por nodo
        conteoR = {i:0 for i in truckRoute} # Conteo de recepciones por nodo

        indices = {truckRoute[i]: i for i in range(len(truckRoute))}

        infactibilidades = 0
        for route in droneRoute:
            conteoL[route[0]] += 1
            conteoR[route[2]] += 1

            if indices[route[2]] < indices[route[0]]: # La ruta del dron debe seguir el sentido de la ruta del camión.
                infactibilidades += 1
            
            # Verificando duranción de los vuelos de los drones, considerando tiempos de espera.
            if route[2] != truckRoute[-1]:
                
                espera = 0 # Tiempo de espera de dron
                se = 0
                for vuelo in droneRoute:
                    if vuelo[0] == route[0]:
                        se +=1
                if route[0] == 0:
                    if tiemposArribo[route[0]] + se + time_d[route[0]][route[1]] + serv_d + time_d[route[1]][route[2]] < tiemposArribo[route[2]]: # El camión llega después
                        espera += tiemposArribo[route[2]] - (tiemposArribo[route[0]] + se + time_d[route[0]][route[1]] + serv_d + time_d[route[1]][route[2]])

                else:
                    if tiemposArribo[route[0]] + se + serv_t + time_d[route[0]][route[1]] + serv_d + time_d[route[1]][route[2]] < tiemposArribo[route[2]]: # El camión llega después
                        espera += tiemposArribo[route[2]] - (tiemposArribo[route[0]] + se + serv_t + time_d[route[0]][route[1]] + serv_d + time_d[route[1]][route[2]])

                if time_d[route[0]][route[1]] + serv_d + time_d[route[1]][route[2]] + espera > endurance:  #El vuelo del dron es factible por el tiempo de la distancia recorrida, sin considerar espera
                        infactibilidades += 1

                elif time_d[route[0]][route[1]] + serv_d + time_d[route[1]][route[2]] > endurance: # El vuelo del dron, sin considerar la espera, es infactible.
                    infactibilidades += 1
                
        free_drones = nrDrones
        for i in truckRoute:

            if nrDrones < conteoL[i]: # Existen más lanzamientos en un nodo que la cantidad de drones.
                infactibilidades += 1
                
            if nrDrones < conteoR[i]: # Existen más recepciones en un nodo que la cantidad de drones.
                infactibilidades += 1

            # Verificando balance de drones volando.
            if conteoL[i] > conteoR[i]:
                cont = 0 # En caso de que exista un vuelo <i,j,i> y el balance (lanzaminetos-recepciones) es (+), se consideran estos vuelos solo como lanzamientos.
                for j in droneRoute:
                    if j[0] == i and j[2] == i:
                        cont += 1
                free_drones -= (conteoL[i] - conteoR[i] + cont)
            
            elif conteoL[i] < conteoR[i]:
                free_drones +=  conteoR[i] - conteoL[i]
            
            if free_drones < 0: # Hay un desbalance en la cantidad de drones volando.
                infactibilidades += 1
                    
    else:
        infactibilidades = 0

    return infactibilidades

def SA(truckRoute, droneRoute, tiempos, alfa, temp, finalTemp, iterMax_SA, n_SA, penalizacion, tiempoLimite, cytn):

    tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, droneRoute, tpoArribo = tiempos)
    infactibilidades = factibilidad_c(time_d, serv_t, serv_d, se, nrDrones, endurance, truckRoute, droneRoute, tiempos)

    if infactibilidades != 0: # Existen infactibilidades, se penalizará el tiempo de arribo al depósito.. Se dejarán los tiempos de arribo calculados, aunque generen distoriciones en los tiempos de viaje.
            tiempos[truckRoute[-1]] += penalizacion*infactibilidades

    actualTimes = tiempos.copy()
    bestTimes = tiempos.copy()
    best_infactibilidades = infactibilidades
    actual_infactibilidades = infactibilidades
    actual_droneRoute = copy.deepcopy(droneRoute)
    bestDroneRoute = copy.deepcopy(droneRoute)
    
    iter_SA = 1
    ini_SA = time.time()
    while finalTemp <= temp and iter_SA < iterMax_SA:
        
        indice_d = random.randint(0, len(droneRoute) - 1)

        for i in range(n_SA):
            d_candidato = copy.deepcopy(actual_droneRoute)
            
            # Perturbación
            perturbacionDron2(truckRoute, d_candidato, indice_d)

            # Factibilidad
            tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, d_candidato, tpoArribo = tiempos)
            infactibilidades = factibilidad_c(time_d, serv_t, serv_d, se, nrDrones, endurance ,truckRoute, d_candidato, tiempos)

            if infactibilidades != 0: # Existen infactibilidades, se penalizará el tiempo de arribo al depósito.. Se dejarán los tiempos de arribo calculados, aunque generen distoriciones en los tiempos de viaje.
                tiempos[truckRoute[-1]] += penalizacion*infactibilidades

            # Aceptar soluciones
            if actual_infactibilidades == 0 and infactibilidades == 0:
                if tiempos[truckRoute[-1]] < actualTimes[truckRoute[-1]]:
                    actualTimes = tiempos.copy()
                    actual_droneRoute = copy.deepcopy(d_candidato)
                    actual_infactibilidades = infactibilidades

                    if best_infactibilidades == 0 and infactibilidades == 0:
                        if tiempos[truckRoute[-1]] < bestTimes[truckRoute[-1]]:
                            bestTimes = tiempos.copy()
                            bestDroneRoute = copy.deepcopy(d_candidato)
                            best_infactibilidades = infactibilidades

                    elif best_infactibilidades > 0 and infactibilidades == 0:
                            bestTimes = tiempos.copy()
                            bestDroneRoute = copy.deepcopy(d_candidato)
                            best_infactibilidades = infactibilidades

                elif random.uniform(0, 1) < math.exp(-abs(tiempos[truckRoute[-1]] - actualTimes[truckRoute[-1]]) / temp):
                    actualTimes = tiempos.copy()
                    actual_droneRoute = copy.deepcopy(d_candidato)
                    actual_infactibilidades = infactibilidades

            elif actual_infactibilidades > 0 and infactibilidades == 0:

                actualTimes = tiempos.copy()
                actual_droneRoute = copy.deepcopy(d_candidato)
                actual_infactibilidades = infactibilidades

                if best_infactibilidades == 0 and infactibilidades == 0:
                    if tiempos[truckRoute[-1]] < bestTimes[truckRoute[-1]]:
                        bestTimes = tiempos.copy()
                        bestDroneRoute = copy.deepcopy(d_candidato)
                        best_infactibilidades = infactibilidades

                elif best_infactibilidades > 0 and infactibilidades == 0:
                    bestTimes = tiempos.copy()
                    bestDroneRoute = copy.deepcopy(d_candidato)
                    best_infactibilidades = infactibilidades

            elif actual_infactibilidades > 0 and infactibilidades > 0:

                if tiempos[truckRoute[-1]] < actualTimes[truckRoute[-1]]:
                    actualTimes = tiempos.copy()
                    actual_droneRoute = copy.deepcopy(d_candidato)
                    actual_infactibilidades = infactibilidades

                    if best_infactibilidades > 0 and infactibilidades > 0:
                        if tiempos[truckRoute[-1]] < bestTimes[truckRoute[-1]]:
                            bestTimes = tiempos.copy()
                            bestDroneRoute = copy.deepcopy(d_candidato)
                            best_infactibilidades = infactibilidades
                                
                elif random.uniform(0, 1) < math.exp(-abs(tiempos[truckRoute[-1]] - actualTimes[truckRoute[-1]]) / temp):
                    actualTimes = tiempos.copy()
                    actual_droneRoute = copy.deepcopy(d_candidato)
                    actual_infactibilidades = infactibilidades

            if time.time() - ini_SA >= tiempoLimite/10:
                break
        if time.time() - ini_SA >= tiempoLimite/10:
            break

        temp *= alfa
        iter_SA += 1

    return bestDroneRoute, bestTimes, best_infactibilidades

def VNS(ruta_lkh, tiempoLimite, iterMax_VNS, K_VNS, mod1, mod2, p_VNS, q_VNS, alfa, temp, finalTemp, iterMax_SA, n_SA, penalizacion, cytn, rutaCamion= '', rutaDron= '', nodosDron= ''):      

    inicioVNS = time.time() # Tiempo de inicio

    # Solución inicial
    if type(rutaCamion) == str:
        truckRoute, droneRoute, droneNodes, tiempos, infactibilidades  = sol_inicial(ruta_lkh, cytn) #La solución inicial es la misma para todas las semillas. Tal vez elegir al azar el vuelo que se acepta. Sin cambiar la ruta del camión

    elif type(rutaCamion) != str():
        truckRoute, droneRoute, droneNodes = rutaCamion, rutaDron, nodosDron
        tiempos = {}
        tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, droneRoute, tpoArribo = tiempos)
        infactibilidades = factibilidad_c(time_d, serv_t, serv_d, se, nrDrones, endurance ,truckRoute, droneRoute, tiempos)

        if infactibilidades != 0: # Existen infactibilidades, se penalizará el tiempo de arribo al depósito.. Se dejarán los tiempos de arribo calculados, aunque generen distoriciones en los tiempos de viaje.
            tiempos[truckRoute[-1]] += penalizacion*infactibilidades

    ### Inicial: soluciones y costos. ###
    best_truckRoute = truckRoute.copy()  
    best_droneRoute = copy.deepcopy(droneRoute)
    best_droneNodes = droneNodes.copy()
    best_time = tiempos.copy()
    best_infactibilidades = infactibilidades

    flag_mod = False # Despues de ejecutar el modulo 1, se ejecuta el modulo 2 inmediatamente.

    ### Iteraciones principales ###
    for i in range(iterMax_VNS):

        if time.time() - inicioVNS >= tiempoLimite:
            break

        if i%mod1 == 0:
            # Candidato
            candidateTruckRoute = truckRoute.copy()
            candidateDroneRoute = copy.deepcopy(droneRoute)
            candidateDroneNodes = droneNodes.copy()

            tiempos = {}
            tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, droneRoute, tpoArribo = tiempos)
            fact = factibilidad_c(time_d, serv_t, serv_d, se, nrDrones, endurance, truckRoute, droneRoute, tiempos)

            a = random.random()

            for k in range(K_VNS): # A la primera mejora encontrada se detiene
                
                RCST(candidateTruckRoute, candidateDroneRoute, candidateDroneNodes)

                candidato_tiempos = {}
                tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, candidateTruckRoute, candidateDroneRoute, tpoArribo = candidato_tiempos)
                candidato_fact = factibilidad_c(time_d, serv_t, serv_d, se, nrDrones, endurance, candidateTruckRoute, candidateDroneRoute, candidato_tiempos)
                
                if candidato_tiempos[candidateTruckRoute[-1]] < tiempos[truckRoute[-1]] and candidato_fact == 0:
                    truckRoute = candidateTruckRoute.copy()
                    droneRoute = copy.deepcopy(candidateDroneRoute)
                    droneNodes = candidateDroneNodes.copy()
                    tiempos = candidato_tiempos.copy()
                    fact = candidato_fact
                    break # Se detiene, mejora encontrada

                if a <= p_VNS:
                    ICST(candidateTruckRoute, candidateDroneRoute, candidateDroneNodes)
                
                else:
                    TRA(candidateTruckRoute, candidateDroneRoute, candidateDroneNodes)

                candidato_tiempos = {}
                tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, candidateTruckRoute, candidateDroneRoute, tpoArribo = candidato_tiempos)
                candidato_fact = factibilidad_c(time_d, serv_t, serv_d, se, nrDrones, endurance, candidateTruckRoute, candidateDroneRoute, candidato_tiempos)

                if candidato_fact > 0:
                    candidato_tiempos[candidateTruckRoute[-1]] += penalizacion*candidato_fact

                if candidato_tiempos[candidateTruckRoute[-1]] < tiempos[truckRoute[-1]] and candidato_fact == 0:
                    truckRoute = candidateTruckRoute.copy()
                    droneRoute = copy.deepcopy(candidateDroneRoute)
                    droneNodes = candidateDroneNodes.copy()
                    tiempos = candidato_tiempos.copy()
                    fact = candidato_fact
                    break # Se detiene, mejora encontrada
            
            # Actualizar mejor solución
            if best_infactibilidades == 0 and fact == 0:
                if tiempos[truckRoute[-1]] < best_time[best_truckRoute[-1]]:
                    best_truckRoute = truckRoute.copy()
                    best_droneRoute = copy.deepcopy(droneRoute)
                    best_droneNodes = droneNodes.copy()
                    best_time = tiempos.copy()
                    best_infactibilidades = fact

            elif best_infactibilidades > 0 and fact == 0:                       
                    best_truckRoute = truckRoute.copy()
                    best_droneRoute = copy.deepcopy(droneRoute)
                    best_droneNodes = droneNodes.copy()
                    best_time = tiempos.copy()
                    best_infactibilidades = fact

            elif best_infactibilidades > 0 and fact > 0:
                if tiempos[truckRoute[-1]] < best_time[best_truckRoute[-1]]:
                    best_truckRoute = truckRoute.copy()
                    best_droneRoute = copy.deepcopy(droneRoute)
                    best_droneNodes = droneNodes.copy()
                    best_time = tiempos.copy()
                    best_infactibilidades = fact
            
            flag_mod = True # Para ejecutar inmediatamente el SA


        if i%mod2 == 0 or flag_mod == True:

            candidate_time_truckRoute = {}
            tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, tpoArribo = candidate_time_truckRoute)
            
            # SOLUCION CON SA
            if len(droneNodes) != 0:
                droneRoute, tiempos2, infactibilidades2 = SA(truckRoute, droneRoute, candidate_time_truckRoute, alfa, temp, finalTemp, iterMax_SA, n_SA, penalizacion, tiempoLimite, cytn)
                
            else: # tsp solo camion
                infactibilidades2 = 0
                droneRoute = []
                tiempos2 = {}
                tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, truckRoute, tpoArribo = tiempos2)
                
            if best_infactibilidades == 0 and infactibilidades2 == 0:
                if tiempos2[truckRoute[-1]] < best_time[best_truckRoute[-1]]:
                    best_truckRoute = truckRoute.copy()
                    best_droneRoute = copy.deepcopy(droneRoute)
                    best_droneNodes = droneNodes.copy()
                    best_time = tiempos2.copy()
                    best_infactibilidades = infactibilidades2

            elif best_infactibilidades > 0 and infactibilidades2 == 0:                                 
                best_truckRoute = truckRoute.copy()
                best_droneRoute = copy.deepcopy(droneRoute)
                best_droneNodes = droneNodes.copy()
                best_time = tiempos2.copy()
                best_infactibilidades = infactibilidades2


            elif best_infactibilidades > 0 and infactibilidades2 > 0:
                if tiempos2[truckRoute[-1]] < best_time[best_truckRoute[-1]]:
                    best_truckRoute = truckRoute.copy()
                    best_droneRoute = copy.deepcopy(droneRoute)
                    best_droneNodes = droneNodes.copy()
                    best_time = tiempos2.copy()
                    best_infactibilidades = infactibilidades2

            flag_mod = False

    return best_time[best_truckRoute[-1]], best_truckRoute, best_droneRoute, best_droneNodes , best_infactibilidades #, tipo_sol

def matheuristica(tiempoLimMatheuristica, tiempoLimiteVNS, seed, ruta_lkh, tiempoLimiteMdl, iterMax_VNS, K_VNS, mod1, mod2, p_VNS, q_VNS, alfa, temp, finalTemp, iterMax_SA, n_SA, penalizacion, cytn):

    iniMatheuristica = time.time()
    mejor_costo, mejor_truckRoute, mejor_droneRoute, mejor_droneNodes, mejor_factibilidades  = VNS(ruta_lkh, tiempoLimiteVNS, iterMax_VNS, K_VNS, mod1, mod2, p_VNS, q_VNS, alfa, temp, finalTemp, iterMax_SA, n_SA, penalizacion, cytn)
    
    flag = True
    while time.time() - iniMatheuristica <= tiempoLimMatheuristica:

        if (tiempoLimMatheuristica - (time.time() - iniMatheuristica)) - tiempoLimiteMdl >= 0 and flag == True and len(mejor_droneRoute) != 0:
            t_TSP = {}
            tiemposArribo_c(time_t, time_d, serv_t, serv_d, se, mejor_truckRoute, tpoArribo = t_TSP)
        
            droneRoute_mdl, droneNodes_mdl, tiempos_mdl, factibilidades_mdl = modelo_Gurobi.modelo_FDTSP_reducido(tiempoLimiteMdl, seed, coords= coords ,time_t= time_t, time_d= time_d, truckRoute= mejor_truckRoute, m= nrDrones, endurance = endurance, se = se, serv_t = serv_t, serv_d = serv_d, tiemposTSP= t_TSP, droneRoute= mejor_droneRoute)
            
            if mejor_factibilidades == 0 and factibilidades_mdl == 0:
                if tiempos_mdl[mejor_truckRoute[-1]] < mejor_costo:
                    mejor_droneRoute = copy.deepcopy(droneRoute_mdl)
                    mejor_droneNodes = droneNodes_mdl.copy()
                    mejor_costo = tiempos_mdl[mejor_truckRoute[-1]]
                    mejor_factibilidades = factibilidades_mdl

        if (tiempoLimMatheuristica - (time.time() - iniMatheuristica)) - tiempoLimiteVNS >= 0:
            costo_vns, truckRoute_vns, droneRoute_vns, droneNodes_vns, factibilidades_vns = VNS(ruta_lkh, tiempoLimiteVNS, iterMax_VNS, K_VNS, mod1, mod2, p_VNS, q_VNS, alfa, temp, finalTemp, iterMax_SA, n_SA, penalizacion, cytn, rutaCamion= mejor_truckRoute, rutaDron= mejor_droneRoute, nodosDron= mejor_droneNodes)

            if mejor_factibilidades == 0 and factibilidades_vns == 0:
                if costo_vns < mejor_costo:
                    mejor_truckRoute = truckRoute_vns.copy()
                    mejor_droneRoute = copy.deepcopy(droneRoute_vns)
                    mejor_droneNodes = droneNodes_vns.copy()
                    mejor_costo = costo_vns
                    mejor_factibilidades = factibilidades_vns

    return mejor_costo, mejor_truckRoute, mejor_droneRoute

def main(inst, iterMax_VNS, K_VNS, mod1, mod2, p_VNS, q_VNS, alfa, temp, finalTemp, iterMax_SA, n_SA, penalizacion, cytn):
    global nrNodos, endurance, droneSpeed, truckSpeed, coords, se, serv_t, serv_d, nrDrones, time_t, time_d, seed, grid

    instances_path = '../instancias' # './instancias_modelo'

    ruta_lkh = 'LKH-3.0.7/LKH' # Ruta del archivo ejecutable LKH 

    nrDrones = 2 # Nr. de drones
    nrSeeds = 10 # Nr. de semillas

    drones = [i for i in range(2, nrDrones + 1)]
    seeds = [i for i in range(1, nrSeeds + 1)]

    se = 1          # Tiempo de set up para el lanzamiento de un dron
    serv_t = 0.5    # Tiempo de servicio del camion
    serv_d = 0.5    # Tiempo de servicio de un drone

    nrNodos, endurance, droneSpeed, truckSpeed, coords = lecturaDatos(inst, instances_path)
    time_t, time_d = tiemposViaje()
    grid = inst[:4]

    # Tiempos límite de ejecucion según cantidad de nodos de la instancia.
    if  nrNodos <= 40:
        tiempoLimiteVNS = 1 # segundos
        tiempoLimiteMdl = 3
        tiempoLimMatheuristica = 10
    elif nrNodos == 60 or nrNodos == 80:
        tiempoLimiteVNS = 10 
        tiempoLimiteMdl = 240 # 150
        tiempoLimMatheuristica = 300
    elif nrNodos == 100 or nrNodos == 150:
        tiempoLimiteVNS = 10 
        tiempoLimiteMdl = 480 # 300
        tiempoLimMatheuristica = 600

    suma = 0
    best_plot_solution = None
    for nrDrones in drones:
        for seed in seeds:
            random.seed(seed)
            np.random.seed(seed)

            inicioTiempo = time.time() 
            costo, truckRoute, droneRoute = matheuristica(tiempoLimMatheuristica, tiempoLimiteVNS, seed, ruta_lkh, tiempoLimiteMdl, iterMax_VNS, K_VNS, mod1, mod2, p_VNS, q_VNS, alfa, temp, finalTemp, iterMax_SA, n_SA, penalizacion, cytn)
            tiempoEjecucion = time.time() - inicioTiempo

            # Keep the best solution across all drone-fleet sizes and seeds so
            # that one representative final route is plotted for the instance.
            if best_plot_solution is None or costo < best_plot_solution["cost"]:
                best_plot_solution = {
                    "cost": costo,
                    "truck_route": truckRoute.copy(),
                    "drone_routes": copy.deepcopy(droneRoute),
                    "number_of_drones": nrDrones,
                    "seed": seed,
                }

            suma += costo

            ### Metricas ###
            t = {}
            espera = tiemposArribo_func(truckRoute= truckRoute, droneRoute= droneRoute, tpoArribo= t, metrica= 'SI')
            waiting_time = espera/ costo

            drone_delivery = len(droneRoute) / (nrNodos - 1)

            # service_level: razon de clientes atendidos en menos de 4 horas (240 min) y el total de clientes.
            conteo = 0
            for i in t:
                if i != 0 and i != nrNodos and t[i] > 240: # Corregir: <
                    conteo += 1
            service_level = conteo / (nrNodos - 1)

            print(inst, nrDrones, seed, costo, tiempoEjecucion, drone_delivery, service_level, waiting_time)

    if best_plot_solution is not None:
        png_path, pdf_path = plot_solution_route(
            coords=coords,
            truck_route=best_plot_solution["truck_route"],
            drone_routes=best_plot_solution["drone_routes"],
            instance_name=inst,
            number_of_drones=best_plot_solution["number_of_drones"],
            seed_value=best_plot_solution["seed"],
            objective_value=best_plot_solution["cost"],
            output_dir="route_plots",
        )
        print("Route plot saved:", png_path)
        print("Vector route plot saved:", pdf_path)

    ######### Registro resultados en archivo.xlsx #########
    # from openpyxl import Workbook
    # wb = Workbook()
    # wb.remove(wb.active)

    # for i in resultados: # cada instancia
    #     datos = [i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7], i[8]]
    #     try:
    #         ws = wb['Metaheuristica']
    #     except KeyError:
    #         ws = wb.create_sheet('Metaheuristica')
    #         ws.append(['Instancia','nrDrones', 'semilla', 'Costo algoritmo', 'Tiempo ejecucion', 'Costo mdl medio', 'RunTime mdl medio', 'Costo mdl final', 'RunTime mdl final'])
    #     ws.append(datos)

    # wb.save('./resultados.xlsx')

if __name__=="__main__":
    argv = sys.argv[1:]
    try:
        opts, args = getopt.getopt(argv,"a:b:c:d:e:f:g:h:i:j:k:l:m:n:o:")
        # print(argv, len (opts)) 
        if len (opts) < 15:
            print('.\FDTSP_Algorithm.py -a -b -c -d -e -f -g -h -i -j -k -l -m -n -o')
            sys.exit(2)
    except getopt.GetoptError:
        print('.\FDTSP_Algorithm.py -a -b -c -d -e -f -g -h -i -j -k -l -m -n -o')        
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-a':
            iterMax_VNS = int(arg)
        elif opt == '-b':
            K_VNS = int(arg)
        elif opt == '-c':
            mod1 = int(arg)
        elif opt == '-d':
            mod2 = int(arg)
        elif opt == '-e':
            p_VNS = float(arg)
        elif opt == '-f':
            q_VNS = float(arg)
        elif opt == '-g':
            alfa = float(arg)
        elif opt == '-h':
            temp = float(arg)
        elif opt == '-i':
            finalTemp = float(arg)
        elif opt == '-j':
            iterMax_SA = int(arg)
        elif opt == '-k':
            n_SA = int(arg)
        elif opt == '-l':
            penalizacion = int(arg)
        elif opt == '-m':
            cytn = int(arg) 
        elif opt == '-n':
            inst = arg
        elif opt == 'o':
            r_VNS = arg

    # print('instancia:', inst)
    # print('a:', iterMax_VNS)
    # print('b:', K_VNS)
    # print('c:', mod1)
    # print('d:', mod2)
    # print('e:', p_VNS)
    # print('f:', q_VNS)
    # print('g:', alfa)
    # print('h:', temp)
    # print('i:', finalTemp)
    # print('j:', iterMax_SA)
    # print('k:', n_SA)
    # print('l:', p_dron)
    # print('m:', q_dron)
    # print('o:', penalizacion)
    # print('p: ', cytn)
    # print('q: ', inst)
    
    file = open(inst, "r")
    for i in file:
        instancia = i[:-1]
        main(instancia, iterMax_VNS, K_VNS, mod1, mod2, p_VNS, q_VNS, alfa, temp, finalTemp, iterMax_SA, n_SA, penalizacion, cytn)
        # break
    file.close()
    



